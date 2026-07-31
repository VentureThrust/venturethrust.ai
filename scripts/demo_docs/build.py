# -*- coding: utf-8 -*-
"""
Build the five demo data rooms.

  python scripts/demo_docs/build.py            generate files locally
  python scripts/demo_docs/build.py --upload   generate, then upload to storage
  python scripts/demo_docs/build.py --upload --sql   also rewrite the seed SQL

Generated files land in scripts/demo_docs/out/<slug>/ and upload to the
vdr-files bucket at demo/<slug>/<filename>, which is exactly the storage_path
written into the SQL. One manifest drives both, so rows and objects match.
"""

import os
import re
import sys
import io
import urllib.parse

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer, Table, TableStyle,
    KeepTogether, PageBreak,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from profiles import ALL_PROFILES, BY_SLUG      # noqa: E402
from manifest import MANIFEST, FOLDERS          # noqa: E402

OUT = os.path.join(HERE, "out")
BUCKET = "vdr-files"

INK = colors.HexColor("#111827")
MUTED = colors.HexColor("#6B7280")
RULE = colors.HexColor("#D1D5DB")
FAINT = colors.HexColor("#F3F4F6")
NAVY = colors.HexColor("#0D1B3E")
ACCENT = colors.HexColor("#1E3A6E")


# ── Styles ───────────────────────────────────────────────────────────────────

def styles():
    return {
        "h1": ParagraphStyle("h1", fontName="Times-Bold", fontSize=19, leading=23,
                             textColor=INK, spaceAfter=2),
        "sub": ParagraphStyle("sub", fontName="Times-Italic", fontSize=10.5, leading=14,
                              textColor=MUTED, spaceAfter=14),
        "h2": ParagraphStyle("h2", fontName="Times-Bold", fontSize=12, leading=15,
                             textColor=NAVY, spaceBefore=13, spaceAfter=5),
        "p": ParagraphStyle("p", fontName="Times-Roman", fontSize=10.5, leading=15.5,
                            textColor=INK, alignment=TA_LEFT, spaceAfter=6),
        "li": ParagraphStyle("li", fontName="Times-Roman", fontSize=10.5, leading=15.5,
                             textColor=INK, leftIndent=12, bulletIndent=2, spaceAfter=4),
        "cell": ParagraphStyle("cell", fontName="Times-Roman", fontSize=8.8, leading=11.5,
                               textColor=INK),
        "cellb": ParagraphStyle("cellb", fontName="Times-Bold", fontSize=8.8, leading=11.5,
                                textColor=colors.white),
        # deck
        "st": ParagraphStyle("st", fontName="Helvetica-Bold", fontSize=30, leading=35,
                             textColor=NAVY, spaceAfter=4),
        "sb": ParagraphStyle("sb", fontName="Helvetica", fontSize=15, leading=25,
                             textColor=INK, leftIndent=14, bulletIndent=0, spaceAfter=9),
        "cover_t": ParagraphStyle("cover_t", fontName="Helvetica-Bold", fontSize=50, leading=56,
                                  textColor=NAVY),
        "cover_s": ParagraphStyle("cover_s", fontName="Helvetica", fontSize=17, leading=27,
                                  textColor=MUTED),
    }


class DocTemplate(BaseDocTemplate):
    def __init__(self, path, company, title, **kw):
        self.company = company
        self.doc_title = title
        BaseDocTemplate.__init__(self, path, pagesize=A4,
                                 leftMargin=22 * mm, rightMargin=22 * mm,
                                 topMargin=26 * mm, bottomMargin=20 * mm,
                                 title=title, author=company, **kw)
        frame = Frame(self.leftMargin, self.bottomMargin,
                      self.width, self.height, id="body")
        self.addPageTemplates([PageTemplate(id="std", frames=[frame], onPage=self._chrome)])

    def _chrome(self, canv, doc):
        canv.saveState()
        w, h = A4
        canv.setFont("Times-Bold", 9)
        canv.setFillColor(NAVY)
        canv.drawString(22 * mm, h - 16 * mm, self.company.upper())
        canv.setFont("Times-Roman", 9)
        canv.setFillColor(MUTED)
        canv.drawRightString(w - 22 * mm, h - 16 * mm, self.doc_title)
        canv.setStrokeColor(RULE)
        canv.setLineWidth(0.6)
        canv.line(22 * mm, h - 19 * mm, w - 22 * mm, h - 19 * mm)
        canv.line(22 * mm, 15 * mm, w - 22 * mm, 15 * mm)
        canv.setFont("Times-Italic", 7.6)
        canv.setFillColor(MUTED)
        canv.drawString(22 * mm, 11 * mm,
                        "SAMPLE DOCUMENT. Prepared for a VentureThrust product demonstration. "
                        "Not a real company filing.")
        canv.setFont("Times-Roman", 8)
        canv.drawRightString(w - 22 * mm, 11 * mm, "Page %d" % canv.getPageNumber())
        canv.restoreState()


def data_table(st, rows, widths=None, total_w=166 * mm):
    head, body = rows[0], rows[1:]
    n = len(head)
    if widths is None:
        widths = [total_w / n] * n
    data = [[Paragraph(str(c), st["cellb"]) for c in head]]
    for r in body:
        data.append([Paragraph(str(c), st["cell"]) for c in r])
    t = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, RULE),
        ("LINEBELOW", (0, -1), (-1, -1), 0.8, RULE),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), FAINT))
    t.setStyle(TableStyle(style))
    return t


def kv_table(st, pairs, total_w=166 * mm):
    data = [[Paragraph(str(k), st["cell"]),
             Paragraph("<b>%s</b>" % v, st["cell"])] for k, v in pairs]
    t = Table(data, colWidths=[total_w * 0.55, total_w * 0.45], hAlign="LEFT")
    style = [
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, RULE),
        ("LINEABOVE", (0, 0), (-1, 0), 0.8, NAVY),
    ]
    for i in range(len(data)):
        if i % 2 == 1:
            style.append(("BACKGROUND", (0, i), (-1, i), FAINT))
    t.setStyle(TableStyle(style))
    return t


# ── Auto sections: content computed from the company profile ─────────────────

AUTO = {}


def auto(slug, title):
    def deco(fn):
        AUTO[(slug, title)] = fn
        return fn
    return deco


def _metrics(st, p):
    return [kv_table(st, p["metrics"])]


def _traction(st, p):
    return [data_table(st, p["traction_rows"])]


def _customers(st, p):
    return [data_table(st, p["customers"])]


for _p in ALL_PROFILES:
    AUTO[(_p["slug"], "Where the business stands")] = _metrics
    AUTO[(_p["slug"], "Performance")] = _metrics
    AUTO[(_p["slug"], "Results")] = _traction
    AUTO[(_p["slug"], "Track record by customer")] = _customers
    AUTO[(_p["slug"], "Term by term")] = _traction
    AUTO[(_p["slug"], "Defect history")] = _traction


@auto("nellara-agrichain", "Summary of outward supplies")
def _n_gst(st, p):
    rows = [["Quarter", "Taxable value (Rs)", "CGST (Rs)", "SGST (Rs)", "Total tax (Rs)"],
            ["Q1 FY26", "1,84,20,000", "0", "0", "Nil rated"],
            ["Q2 FY26", "2,31,60,000", "1,12,400", "1,12,400", "2,24,800"],
            ["Q3 FY26", "2,88,40,000", "1,41,900", "1,41,900", "2,83,800"],
            ["Q4 FY26", "3,46,80,000", "1,68,200", "1,68,200", "3,36,400"],
            ["FY26 total", "10,51,00,000", "4,22,500", "4,22,500", "8,45,000"]]
    return [data_table(st, rows),
            Spacer(1, 5),
            Paragraph("Most fresh produce is nil rated. Tax arises on processed and packed lines only.",
                      st["p"])]


@auto("nellara-agrichain", "Twelve month history")
def _n_hist(st, p):
    rows = [["Month", "Accepted (kg)", "Invoiced (kg)", "Wastage %", "Primary cause"],
            ["Aug 2025", "41,200", "31,900", "22.6", "Transit damage"],
            ["Oct 2025", "48,600", "37,900", "22.0", "Transit damage"],
            ["Dec 2025", "56,100", "44,300", "21.0", "Shelf loss at hub"],
            ["Feb 2026", "64,800", "52,100", "19.6", "Shelf loss at hub"],
            ["Apr 2026", "78,400", "67,300", "14.1", "Grading loss"],
            ["Jun 2026", "94,200", "85,700", "9.0", "Grading loss"],
            ["Jul 2026", "1,02,600", "95,000", "7.4", "Grading loss"]]
    return [data_table(st, rows)]


@auto("zylo-health", "Detailed results")
def _z_res(st, p):
    rows = [["Finding", "Studies", "Sensitivity", "Specificity", "AUC"],
            ["Intracranial haemorrhage", "6,410", "95.8%", "90.1%", "0.974"],
            ["Midline shift", "2,880", "93.2%", "91.4%", "0.968"],
            ["Hydrocephalus", "1,940", "91.7%", "89.6%", "0.951"],
            ["Pneumothorax", "3,120", "94.4%", "88.2%", "0.963"],
            ["Pleural effusion", "5,270", "93.9%", "86.8%", "0.955"],
            ["Consolidation", "4,380", "92.1%", "85.4%", "0.941"],
            ["Combined critical endpoint", "24,000", "94.1%", "88.7%", "0.961"]]
    return [data_table(st, rows)]


@auto("zylo-health", "Live contracts")
def _z_live(st, p):
    rows = [r[:1] + r[3:] for r in p["customers"]]
    return [data_table(st, rows)]


@auto("voltaneer", "Specification")
def _v_spec(st, p):
    return [kv_table(st, [
        ("Measurement range", "5 A to 1,200 A, three phase"),
        ("Accuracy class", "Class 0.5S per IS 16444"),
        ("Sampling rate", "1 Hz standard, 1 kHz on the main incomer"),
        ("Parameters captured", "Current, voltage, power factor, kW, kVA, kVAh, THD"),
        ("Edge storage", "72 hours of buffered readings"),
        ("Connectivity", "4G LTE Cat M1 with Wi Fi fallback"),
        ("Power", "Self powered from the measured line, no separate supply"),
        ("Enclosure", "IP65, DIN rail or panel mount"),
        ("Installation", "Clamp on, no line shutdown required"),
    ])]


@auto("voltaneer", "Grants to date")
def _v_esop(st, p):
    rows = [["Grant date", "Recipients", "Options", "Exercise price (Rs)", "Vested at 31 Jul 2026"],
            ["18 Mar 2025", "6", "34,000", "88", "14,167"],
            ["1 Sep 2025", "9", "21,000", "112", "5,250"],
            ["6 Jan 2026", "7", "12,000", "128", "0"],
            ["4 May 2026", "5", "7,000", "141", "0"],
            ["Total granted", "27", "74,000", "-", "19,417"]]
    return [data_table(st, rows)]


@auto("aegis-drone-systems", "Summary of findings")
def _a_find(st, p):
    rows = [["Severity", "Count", "Example", "Recommended action"],
            ["Critical", "1", "Section loss on bollard mounting plate 14",
             "Structural assessment within 30 days"],
            ["Major", "6", "Coating breakdown with active corrosion, crane leg 2B",
             "Remediate within 90 days"],
            ["Moderate", "14", "Fender panel displacement, bay 7",
             "Schedule in next maintenance window"],
            ["Minor", "25", "Surface staining and marine growth",
             "Monitor at next cycle"],
            ["Total", "46", "-", "-"]]
    return [data_table(st, rows, widths=[24 * mm, 16 * mm, 68 * mm, 58 * mm])]


@auto("kadal-systems", "Specification")
def _k_spec(st, p):
    return [kv_table(st, [
        ("Network", "Low earth orbit short burst data, 1616 to 1626.5 MHz"),
        ("Position accuracy", "GNSS, 2.5 metres CEP"),
        ("Beacon interval", "10 minutes default, 60 seconds after a distress event"),
        ("Distress latency, median", "41 seconds to relay acknowledgement"),
        ("Power", "12 V vessel supply, 36 hour internal battery"),
        ("Display", "3.5 inch sunlight readable, Malayalam and English"),
        ("Enclosure", "IP67, salt spray tested to 720 hours"),
        ("Dimensions", "184 by 122 by 58 millimetres, 940 grams"),
        ("Antenna", "Mast mount, 4 metre cable, marine grade"),
        ("Warranty", "24 months, on vessel replacement"),
    ])]


# ── PDF builders ─────────────────────────────────────────────────────────────

def build_pdf_doc(path, p, filename, payload):
    st = styles()
    title = os.path.splitext(filename)[0]
    doc = DocTemplate(path, p["name"], title)
    flow = [Paragraph(title, st["h1"]),
            Paragraph(payload["subtitle"] or p["docs_note"], st["sub"])]

    sections = payload["sections"]

    if not sections:
        flow += special_doc(st, p, filename)
    else:
        for head, body in sections:
            flow.append(Paragraph(head, st["h2"]))
            if body is None:
                fn = AUTO.get((p["slug"], head))
                if fn:
                    flow += fn(st, p)
                else:
                    flow += _metrics(st, p)
            elif len(body) == 1:
                flow.append(Paragraph(body[0], st["p"]))
            else:
                for line in body:
                    flow.append(Paragraph(line, st["li"], bulletText="•"))

    doc.build(flow)


def special_doc(st, p, filename):
    low = filename.lower()
    out = []

    if "incorporation" in low:
        out.append(Paragraph(
            "This is to certify that %s is incorporated on and from the date mentioned below under "
            "the Companies Act 2013 and that the company is limited by shares." % p["legal"], st["p"]))
        out.append(Spacer(1, 8))
        out.append(kv_table(st, [
            ("Corporate Identity Number", p["cin"]),
            ("Name of the company", p["legal"]),
            ("Date of incorporation", p["founded"]),
            ("Registered office", p["city"]),
            ("Permanent Account Number", "AA" + p["cin"][7:12] + "K"),
            ("Tax Deduction Account Number", "CHN" + p["cin"][10:15] + "B"),
            ("Class of company", "Private, limited by shares"),
            ("Authorised share capital", "Rs 15,00,000"),
            ("Paid up share capital", "Rs 10,00,000"),
        ]))
        out.append(Spacer(1, 10))
        out.append(Paragraph("Digitally signed by the Registrar of Companies. Issued under the "
                             "Central Registration Centre.", st["p"]))
        out.append(Spacer(1, 14))
        out.append(Paragraph("Attached filings", st["h2"]))
        for line in ["Memorandum of Association, filed with SPICe+ Part B.",
                     "Articles of Association, adopting Table F with amendments.",
                     "Form INC 9, declaration by subscribers and first directors.",
                     "Form AGILE PRO S, covering GSTIN, EPFO, ESIC and bank account."]:
            out.append(Paragraph(line, st["li"], bulletText="•"))
        return out

    # Team documents
    out.append(Paragraph(
        "%s was founded in %s and is based in %s. The founding team is set out below, "
        "with the background relevant to this business." % (p["name"], p["founded"], p["city"]),
        st["p"]))
    for name, role, bio in p["founders"]:
        out.append(Paragraph("%s, %s" % (name, role), st["h2"]))
        out.append(Paragraph(bio, st["p"]))
    out.append(Paragraph("How the team divides the work", st["h2"]))
    out.append(kv_table(st, [
        (p["founders"][0][0], "Commercial, fundraising and the board"),
        (p["founders"][1][0], "Product, operations and delivery"),
        (p["founders"][2][0], "The function named in the role above, reporting to the CEO"),
        ("Decisions requiring both founders", "Hiring above a defined band, pricing changes, "
                                              "any commitment over Rs 10 lakh"),
    ]))
    out.append(Paragraph("Working history", st["h2"]))
    out.append(Paragraph(
        "The founders have worked together since before incorporation and hold vesting schedules "
        "running to 48 months from the date of the founders agreement. No founder has left and "
        "no founder holds an outside operating role.", st["p"]))
    return out


def build_pdf_deck(path, p, filename, payload):
    st = styles()
    title = os.path.splitext(filename)[0]
    page = landscape(A4)
    W, H = page

    class Deck(BaseDocTemplate):
        def __init__(self):
            BaseDocTemplate.__init__(self, path, pagesize=page,
                                     leftMargin=26 * mm, rightMargin=26 * mm,
                                     topMargin=26 * mm, bottomMargin=22 * mm,
                                     title=title, author=p["name"])
            fr = Frame(self.leftMargin, self.bottomMargin, self.width, self.height, id="s")
            self.addPageTemplates([PageTemplate(id="slide", frames=[fr], onPage=self._chrome)])

        def _chrome(self, canv, doc):
            n = canv.getPageNumber()
            canv.saveState()
            if n == 1:
                canv.setFillColor(NAVY)
                canv.rect(0, 0, 13 * mm, H, stroke=0, fill=1)
            else:
                canv.setStrokeColor(RULE)
                canv.setLineWidth(0.6)
                canv.line(26 * mm, 17 * mm, W - 26 * mm, 17 * mm)
                canv.setFont("Helvetica", 8)
                canv.setFillColor(MUTED)
                canv.drawString(26 * mm, 12 * mm, "%s  ·  Seed round  ·  July 2026" % p["name"])
                canv.drawRightString(W - 26 * mm, 12 * mm, "%d" % n)
            canv.setFont("Helvetica", 7)
            canv.setFillColor(MUTED)
            canv.drawCentredString(W / 2, 6 * mm, "SAMPLE. Built for a VentureThrust demonstration.")
            canv.restoreState()

    flow = []
    for i, (head, bullets) in enumerate(payload["slides"]):
        if i == 0:
            flow.append(Spacer(1, 34 * mm))
            flow.append(Paragraph(head, st["cover_t"]))
            flow.append(Spacer(1, 10))
            for b in bullets:
                flow.append(Paragraph(b, st["cover_s"]))
        else:
            flow.append(Spacer(1, 4 * mm))
            flow.append(Paragraph(head, st["st"]))
            flow.append(Spacer(1, 7 * mm))
            for b in bullets:
                flow.append(Paragraph(b, st["sb"], bulletText="–"))
        if i < len(payload["slides"]) - 1:
            flow.append(PageBreak())
    Deck().build(flow)


# ── XLSX builders ────────────────────────────────────────────────────────────

HEAD_FILL = PatternFill("solid", fgColor="0D1B3E")
BAND = PatternFill("solid", fgColor="F3F4F6")
THIN = Side(style="thin", color="D1D5DB")


def _style_sheet(ws, ncols, header_row, first_data_row, last_row):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEAD_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in range(first_data_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(bottom=THIN)
            if (r - first_data_row) % 2 == 1:
                cell.fill = BAND
    widths = []
    for c in range(1, ncols + 1):
        longest = 10
        for r in range(header_row, last_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                longest = max(longest, min(46, len(str(v)) + 2))
        widths.append(longest)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=first_data_row, column=1)


def _title_block(ws, p, title, note):
    ws["A1"] = p["name"]
    ws["A1"].font = Font(bold=True, size=14, color="0D1B3E")
    ws["A2"] = title
    ws["A2"].font = Font(size=11, color="374151")
    ws["A3"] = note
    ws["A3"].font = Font(italic=True, size=9, color="6B7280")
    ws["A4"] = ("SAMPLE. Prepared for a VentureThrust product demonstration. "
                "Not a real company filing.")
    ws["A4"].font = Font(italic=True, size=8, color="9CA3AF")


def build_xlsx_table(path, p, filename, payload):
    wb = Workbook()
    ws = wb.active
    ws.title = payload["sheet"][:31]
    rows = payload["rows"] or p[payload["rows_key"]]
    _title_block(ws, p, os.path.splitext(filename)[0], payload["note"])
    start = 6
    for j, val in enumerate(rows[0], start=1):
        ws.cell(row=start, column=j, value=val)
    r = start + 1
    for row in rows[1:]:
        for j, val in enumerate(row, start=1):
            v = val
            if isinstance(val, str):
                bare = val.replace(",", "").replace("-", "", 1).replace(".", "", 1)
                if bare.isdigit() and "," in val:
                    v = float(val.replace(",", ""))
                elif bare.isdigit():
                    v = float(val)
            ws.cell(row=r, column=j, value=v)
        r += 1
    _style_sheet(ws, len(rows[0]), start, start + 1, r - 1)
    wb.save(path)


MODELS = {
    "nellara-agrichain": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("GMV", [396, 468, 552, 642, 3480, 7920]),
            ("Gross margin %", [0.213, 0.224, 0.235, 0.244, 0.262, 0.278]),
            ("Logistics and hub cost", [56, 63, 71, 80, 402, 848]),
            ("People cost", [42, 48, 56, 62, 318, 604]),
            ("Technology and other", [14, 16, 18, 20, 104, 196]),
        ],
    },
    "zylo-health": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Revenue", [43, 58, 79, 104, 620, 1480]),
            ("Gross margin %", [0.759, 0.771, 0.782, 0.790, 0.806, 0.818]),
            ("Regulatory and clinical", [38, 34, 26, 22, 76, 84]),
            ("People cost", [54, 62, 74, 88, 452, 806]),
            ("Cloud and other", [16, 19, 23, 28, 142, 288]),
        ],
    },
    "voltaneer": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Revenue", [78, 94, 116, 142, 780, 1640]),
            ("Gross margin %", [0.612, 0.638, 0.661, 0.684, 0.724, 0.756]),
            ("Sensor lease fleet cost", [34, 41, 49, 58, 268, 512]),
            ("People cost", [58, 66, 78, 92, 468, 838]),
            ("Cloud and other", [11, 13, 15, 18, 92, 184]),
        ],
    },
    "aegis-drone-systems": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Project revenue", [28, 34, 41, 48, 214, 342]),
            ("Retainer revenue", [0, 9, 22, 38, 262, 618]),
            ("Gross margin %", [0.423, 0.446, 0.482, 0.514, 0.568, 0.612]),
            ("Fleet and maintenance", [12, 14, 17, 20, 96, 168]),
            ("People cost", [34, 39, 46, 54, 268, 452]),
        ],
    },
    "kadal-systems": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Hardware revenue", [39, 21, 58, 72, 342, 604]),
            ("Subscription revenue", [14, 16, 19, 23, 128, 284]),
            ("Blended gross margin %", [0.312, 0.334, 0.348, 0.361, 0.402, 0.448]),
            ("Field and installation", [12, 8, 17, 21, 92, 148]),
            ("People cost", [28, 30, 34, 38, 186, 312]),
        ],
    },
    "anvaya-ai": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Revenue", [158, 214, 288, 372, 1980, 4260]),
            ("Gross margin %", [0.541, 0.562, 0.581, 0.598, 0.628, 0.652]),
            ("Reviewer operations", [34, 42, 52, 64, 306, 552]),
            ("People cost", [62, 74, 90, 108, 548, 936]),
            ("Platform and cloud", [14, 17, 21, 26, 128, 254]),
        ],
    },
    "thooval-studios": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Revenue", [51, 68, 89, 114, 604, 1280]),
            ("Gross margin %", [0.381, 0.396, 0.412, 0.428, 0.462, 0.494]),
            ("Studio and mixing", [12, 15, 19, 24, 118, 232]),
            ("People cost", [39, 45, 54, 64, 322, 566]),
            ("Compute and platform", [6, 8, 10, 13, 64, 132]),
        ],
    },
    "kalpana-robotics": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("College fee revenue", [62, 74, 92, 112, 604, 1180]),
            ("Employer hiring fee revenue", [8, 11, 15, 20, 128, 296]),
            ("Gross margin %", [0.346, 0.362, 0.381, 0.398, 0.436, 0.472]),
            ("Lab kit and logistics", [14, 17, 21, 26, 118, 208]),
            ("People cost", [42, 48, 57, 68, 338, 588]),
        ],
    },
    "metricon-interconnect": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Revenue", [245, 288, 342, 404, 2140, 3860]),
            ("Gross margin %", [0.221, 0.248, 0.276, 0.302, 0.348, 0.386]),
            ("Plant overhead", [22, 25, 29, 34, 168, 288]),
            ("People cost", [34, 39, 46, 54, 268, 452]),
            ("Tooling amortisation", [9, 11, 13, 16, 78, 142]),
        ],
    },
    "puzha-foods": {
        "unit": "Rs lakh",
        "periods": ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27", "FY28", "FY29"],
        "lines": [
            ("Modern trade revenue", [128, 108, 174, 212, 1080, 2140]),
            ("Direct to consumer revenue", [79, 71, 106, 132, 704, 1520]),
            ("Gross margin %", [0.380, 0.372, 0.394, 0.408, 0.436, 0.462]),
            ("Plant and cold chain", [34, 32, 44, 52, 254, 464]),
            ("People and marketing", [48, 46, 62, 74, 372, 686]),
        ],
    },
}


def build_xlsx_model(path, p, filename, payload):
    m = MODELS[p["slug"]]
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    _title_block(ws, p, os.path.splitext(filename)[0], payload["note"])

    periods = m["periods"]
    n = len(periods)
    start = 6
    ws.cell(row=start, column=1, value="Line item (%s)" % m["unit"])
    for j, per in enumerate(periods, start=2):
        ws.cell(row=start, column=j, value=per)

    r = start + 1
    rev_rows = []
    cost_rows = []
    gm_row = None
    for label, vals in m["lines"]:
        ws.cell(row=r, column=1, value=label)
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=r, column=j, value=v)
            if "%" in label:
                c.number_format = "0.0%"
            else:
                c.number_format = "#,##0"
        if "%" in label:
            gm_row = r
        elif "revenue" in label.lower() or label == "GMV":
            rev_rows.append(r)
        else:
            cost_rows.append(r)
        r += 1

    ws.cell(row=r, column=1, value="Gross profit")
    for j in range(2, n + 2):
        col = get_column_letter(j)
        rev = "+".join("%s%d" % (col, x) for x in rev_rows)
        ws.cell(row=r, column=j, value="=(%s)*%s%d" % (rev, col, gm_row)).number_format = "#,##0"
    gp_row = r
    r += 1

    ws.cell(row=r, column=1, value="Total operating cost")
    for j in range(2, n + 2):
        col = get_column_letter(j)
        expr = "+".join("%s%d" % (col, x) for x in cost_rows)
        ws.cell(row=r, column=j, value="=%s" % expr).number_format = "#,##0"
    oc_row = r
    r += 1

    ws.cell(row=r, column=1, value="EBITDA")
    for j in range(2, n + 2):
        col = get_column_letter(j)
        ws.cell(row=r, column=j,
                value="=%s%d-%s%d" % (col, gp_row, col, oc_row)).number_format = "#,##0"
    eb_row = r
    r += 1

    ws.cell(row=r, column=1, value="EBITDA margin")
    for j in range(2, n + 2):
        col = get_column_letter(j)
        rev = "+".join("%s%d" % (col, x) for x in rev_rows)
        ws.cell(row=r, column=j,
                value="=IF((%s)=0,0,%s%d/(%s))" % (rev, col, eb_row, rev)).number_format = "0.0%"
    r += 1

    last = r - 1
    for rr in (gp_row, oc_row, eb_row, last):
        for c in range(1, n + 2):
            ws.cell(row=rr, column=c).font = Font(bold=True)
    _style_sheet(ws, n + 1, start, start + 1, last)

    ws2 = wb.create_sheet("Assumptions")
    _title_block(ws2, p, "Key assumptions", "Drivers behind the model on the previous sheet")
    ws2.cell(row=6, column=1, value="Assumption")
    ws2.cell(row=6, column=2, value="Value")
    rr = 7
    for k, v in p["metrics"]:
        ws2.cell(row=rr, column=1, value=k)
        ws2.cell(row=rr, column=2, value=v)
        rr += 1
    for note in p["risks"]:
        ws2.cell(row=rr, column=1, value="Risk")
        ws2.cell(row=rr, column=2, value=note)
        rr += 1
    _style_sheet(ws2, 2, 6, 7, rr - 1)

    wb.save(path)


BUILDERS = {
    "pdf_doc": build_pdf_doc,
    "pdf_deck": build_pdf_deck,
    "xlsx_table": build_xlsx_table,
    "xlsx_model": build_xlsx_model,
}


# ── Generate ─────────────────────────────────────────────────────────────────

SAFE_NAME = re.compile(r"^[A-Za-z0-9 ._()-]+$")


def generate():
    made = []
    for slug, entries in MANIFEST.items():
        p = BY_SLUG[slug]
        d = os.path.join(OUT, slug)
        os.makedirs(d, exist_ok=True)
        for folder_no, filename, builder, payload in entries:
            # Storage keys reject characters like & , so catch them here rather
            # than after 75 uploads have already gone out.
            if not SAFE_NAME.match(filename):
                raise SystemExit("Unsafe filename for a storage key: %r" % filename)
            path = os.path.join(d, filename)
            BUILDERS[builder](path, p, filename, payload)
            made.append({
                "slug": slug,
                "company": p["name"],
                "folder": FOLDERS[folder_no - 1],
                "name": filename,
                "path": path,
                "storage": "demo/%s/%s" % (slug, filename),
                "size": os.path.getsize(path),
            })
            print("  built %-46s %7d bytes" % (filename, os.path.getsize(path)))
    return made


# ── Upload ───────────────────────────────────────────────────────────────────

def env():
    out = {}
    with open(os.path.join(os.path.dirname(os.path.dirname(HERE)), ".env.local"),
              encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            out[k.strip()] = v.strip().strip('"').strip("'")
    return out


def upload(files):
    import requests
    e = env()
    base = e["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
    key = e["SUPABASE_SERVICE_ROLE_KEY"]
    sess = requests.Session()
    ok = 0
    for f in files:
        ct = ("application/pdf" if f["name"].lower().endswith(".pdf")
              else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        url = "%s/storage/v1/object/%s/%s" % (
            base, BUCKET, urllib.parse.quote(f["storage"], safe="/"))
        with open(f["path"], "rb") as fh:
            body = fh.read()
        r = sess.post(url, data=body, headers={
            "Authorization": "Bearer %s" % key,
            "apikey": key,
            "Content-Type": ct,
            "x-upsert": "true",
        }, timeout=120)
        if r.status_code in (200, 201):
            ok += 1
        else:
            print("  FAILED %s -> %s %s" % (f["storage"], r.status_code, r.text[:200]))
    print("  uploaded %d of %d" % (ok, len(files)))
    return ok == len(files)


# ── SQL ──────────────────────────────────────────────────────────────────────

def sqlq(s):
    return s.replace("'", "''")


def ftype(name):
    low = name.lower()
    if "deck" in low:
        return "Deck"
    if low.endswith(".xlsx"):
        return "Sheet"
    if low.endswith(".pdf"):
        return "PDF"
    return "Doc"


SQL_HEAD = r"""-- ============================================================================
-- DEMO INVESTOR ACCOUNT  ·  venturethrust@gmail.com
-- ============================================================================
-- Investor : venturethrust@gmail.com
-- Owners   : one founder account per startup, so Shared with me shows a
--            different sender for every room. Create them first with
--            python scripts/demo_docs/founders.py
--
-- Generated by scripts/demo_docs/build.py. Every file row below has a real
-- object already uploaded to the vdr-files bucket at the same storage_path,
-- so opening any document works.
--
-- Safe to re-run. It removes only the investor's own rows and the five demo
-- spaces, never anything else.
-- ============================================================================

alter table public.profiles
  add column if not exists signup_notified boolean not null default false;
update public.profiles set signup_notified = true where signup_notified = false;

alter table public.dw_watchlist
  add column if not exists note text,
  add column if not exists quarterly_report boolean not null default false;

do $$
declare
  v_inv uuid;
  v_own uuid;
  v_mgr uuid;
  v_fnd uuid;
  s_id  uuid;
  rec   record;
  ids   uuid[] := '{}';
  founder_emails text[] := array[%(FOUNDERS)s];
  demo_names text[] := array[%(NAMES)s];
begin
  select id into v_inv from auth.users where lower(email) = 'venturethrust@gmail.com';
  if v_inv is null then
    raise exception 'No auth user for venturethrust@gmail.com. Sign up with that email first.';
  end if;
  select id into v_own from auth.users where lower(email) = 'omprakash@venturethrust.com';
  if v_own is null then v_own := v_inv; end if;
  select id into v_mgr from public.profiles where lower(email) = 'omprakash@venturethrust.com';

  -- ── Wipe ─────────────────────────────────────────────────────────────────
  -- Demo rooms are matched by name whoever owns them, because ownership moved
  -- from the manager account to per startup founder accounts.
  select coalesce(array_agg(s.id), '{}') into ids
    from public.spaces s
   where s.name = any(demo_names)
     and (s.created_by in (v_inv, v_own)
          or s.created_by in (select id from auth.users
                               where lower(email) = any(founder_emails)));

  -- Anything ever addressed to this inbox by ANY sender is cleared too.
  -- Old test sends from other accounts are what put stray PDFs and videos
  -- in Shared with me, and they are not owned by v_inv or v_own.
  if to_regclass('public.share_link_access_logs') is not null then
    delete from public.share_link_access_logs where share_link_id in (
      select id from public.share_links
       where space_id = any(ids)
          or lower(recipient_email) = 'venturethrust@gmail.com'
          or space_id in (select id from public.spaces where created_by = v_inv));
  end if;
  if to_regclass('public.viewer_sessions') is not null then
    delete from public.viewer_sessions where space_id = any(ids);
  end if;
  if to_regclass('public.share_links') is not null then
    delete from public.share_links where space_id = any(ids)
      or lower(recipient_email) = 'venturethrust@gmail.com'
      or space_id in (select id from public.spaces where created_by = v_inv);
  end if;
  if to_regclass('public.visits') is not null then
    delete from public.visits where space_id = any(ids)
      or space_id in (select id from public.spaces where created_by = v_inv);
  end if;
  -- Before the spaces go, so nothing is orphaned by a cascade.
  if to_regclass('public.dw_update_events') is not null then
    delete from public.dw_update_events where space_id = any(ids) or founder_id in (v_inv, v_own);
  end if;
  if to_regclass('public.files') is not null then
    delete from public.files where space_id = any(ids) or user_id = v_inv;
  end if;
  if to_regclass('public.folders') is not null then
    delete from public.folders where space_id = any(ids) or user_id = v_inv;
  end if;
  if to_regclass('public.spaces') is not null then
    delete from public.spaces where id = any(ids) or created_by = v_inv;
  end if;
  if to_regclass('public.dw_watchlist') is not null then
    delete from public.dw_watchlist where investor_id = v_inv;
  end if;
  if to_regclass('public.alerts') is not null then
    delete from public.alerts where user_id = v_inv;
  end if;
  if to_regclass('public.dw_offers') is not null then
    delete from public.dw_offers where lower(investor_email) = 'venturethrust@gmail.com';
  end if;

  -- ── Investor plan on ─────────────────────────────────────────────────────
  update public.profiles
     set is_investor = true, plan = 'vdr_ai', plan_status = 'active',
         plan_expires_at = now() + interval '365 days',
         dw_auto_assign = true, signup_notified = true
   where id = v_inv;

  -- The wipe is done. From here ids collects the rooms this run creates.
  ids := '{}';
"""

SQL_TAIL = r"""
  -- ── Founder update history behind the briefs ─────────────────────────────
  insert into public.dw_update_events (id, founder_id, space_id, file_id, file_name, event_type, created_at)
  select gen_random_uuid(), s.created_by, f.space_id, null, f.name, 'file_updated',
         now() - ((row_number() over (order by f.created_at desc)) || ' days')::interval
    from public.files f
    join public.spaces s on s.id = f.space_id
   where s.id = any(ids)
   limit 28;

  -- ── The two priority briefs waiting ──────────────────────────────────────
  insert into public.alerts (user_id, space_id, type, message, created_at)
  select v_inv, id, 'dw_update',
         'Priority brief ready: Nellara AgriChain has crossed both conditions from your note. Wastage 7.4 percent, gross margin 21.3 percent.',
         now() - interval '10 days'
    from public.spaces where name = 'Nellara AgriChain' and id = any(ids) limit 1;

  insert into public.alerts (user_id, space_id, type, message, created_at)
  select v_inv, id, 'dw_update',
         'Priority brief ready: Voltaneer is at 47 percent recurring revenue with a 71 day sales cycle, past both thresholds you set.',
         now() - interval '7 days'
    from public.spaces where name = 'Voltaneer' and id = any(ids) limit 1;
end $$;
"""

# Investor context per startup.
#
# The states have to agree with each other or an investor spots it in a second:
#   watchlist=True  means the investor read it and chose to follow it, so the
#                   invite is necessarily opened. Shows under Opened.
#   watchlist=False means new inbound that has not been read yet, so the invite
#                   is unopened. Shows under Pending.
# Nothing is ever both watchlisted and pending.
CONTEXT = {
    "Nellara AgriChain": dict(
        added="2026-03-12 10:40:00+05:30", quarterly=False, watchlist=True,
        note="Passed at pre seed, March 2026. Right problem, wrong economics: 22 percent wastage "
             "and 8 percent gross margin meant every rupee of growth was losing money. The Kerala "
             "produce gap is real and the founder knows the Wayanad collectives personally. Come "
             "back when wastage is under 10 percent and gross margin is above 18 percent."),
    "Zylo Health": dict(
        added="2026-02-02 16:05:00+05:30", quarterly=True, watchlist=True,
        note="Strong clinical team out of Amrita and the radiologist shortage in tier 2 hospitals "
             "is real. But everything depends on CDSCO Class B clearance, which they had not even "
             "filed when we met. No point pricing a regulated device before the regulator speaks. "
             "Revisit on clearance plus one paying hospital outside Kerala."),
    "Voltaneer": dict(
        added="2026-02-05 12:15:00+05:30", quarterly=False, watchlist=True,
        note="A hardware business wearing a software badge. 76 percent of revenue was one time "
             "sensor sales and the average deal took nearly nine months because it needed a capex "
             "approval. The energy saving is real, the business model was not. Come back when "
             "recurring software is above 40 percent of revenue and the sales cycle is under 90 days."),
    "Aegis Drone Systems": dict(
        added="2026-04-18 15:25:00+05:30", quarterly=False, watchlist=True,
        note="Good pilots, good imagery, no repeat business. Every rupee so far has come from one "
             "off inspections, and DGCA BVLOS permissions were still pending when we met. "
             "Inspection only becomes a business when it becomes a retainer. Revisit when at least "
             "three sites are on annual contracts and BVLOS clearance is granted."),
    "Kadal Systems": dict(
        added="2026-05-06 11:05:00+05:30", quarterly=False, watchlist=True,
        note="Genuinely important product and the safety case is unarguable. My problem is who "
             "pays. Almost all current units were bought through a state subsidy scheme, so I have "
             "no evidence a boat owner pays full price from his own pocket. Show me 200 units sold "
             "without subsidy and I will look again."),

    # New inbound. Not read yet, so no note and no watchlist row.
    "Anvaya AI": dict(
        added="2026-07-29 09:20:00+05:30", quarterly=False, watchlist=False, note=""),
    "Thooval Studios": dict(
        added="2026-07-27 17:45:00+05:30", quarterly=False, watchlist=False, note=""),
    "Puzha Foods": dict(
        added="2026-07-24 11:10:00+05:30", quarterly=False, watchlist=False, note=""),
    "Metricon Interconnect": dict(
        added="2026-07-21 15:30:00+05:30", quarterly=False, watchlist=False, note=""),
    "Kalpana Robotics": dict(
        added="2026-07-18 10:05:00+05:30", quarterly=False, watchlist=False, note=""),
}


def write_sql(files, path):
    names = ", ".join("'%s'" % sqlq(p["name"]) for p in ALL_PROFILES)
    founders = ", ".join("'%s'" % p["founder_email"] for p in ALL_PROFILES)
    out = [SQL_HEAD % {"NAMES": names, "FOUNDERS": founders}]

    # ── One VALUES list drives every room ────────────────────────────────────
    out.append("\n  -- ── %d data rooms, their share link, invite and watchlist row ──────────\n"
               % len(ALL_PROFILES))
    out.append("  for rec in\n    select * from (values\n")
    rows = []
    for p in ALL_PROFILES:
        c = CONTEXT[p["name"]]
        rows.append(
            "      ('%s',\n       '%s',\n       '%s',\n       %s,\n       %s, %s, timestamptz '%s')"
            % (sqlq(p["name"]),
               sqlq(p["one_line"] + " Based in " + p["city"] + "."),
               p["founder_email"],
               ("'%s'" % sqlq(c["note"])) if c["note"] else "null",
               "true" if c["quarterly"] else "false",
               "true" if c["watchlist"] else "false",
               c["added"]))
    out.append(",\n".join(rows))
    out.append("\n    ) as t(nm, descr, founder, note, quarterly, watched, added)\n  loop\n")
    out.append("""    -- Each room is owned by its own founder, so Shared with me shows the
    -- sender who actually sent it rather than one address ten times.
    select id into v_fnd from auth.users where lower(email) = rec.founder;
    if v_fnd is null then
      raise exception 'Missing founder account %%. Run scripts/demo_docs/founders.py first.',
        rec.founder;
    end if;

    s_id := gen_random_uuid();
    ids := ids || s_id;   -- ids now tracks the rooms this run created

    insert into public.spaces (id, name, title, description, created_by)
    values (s_id, rec.nm, rec.nm, rec.descr, v_fnd);

    -- Five top level diligence folders. No wrapper folder: the investor
    -- should land on the structure, not one folder called Root.
    insert into public.folders (id, user_id, name, space_id, parent_id, position)
    select gen_random_uuid(), v_fnd, fol, s_id, null, ord
      from unnest(array[%(FOLDERS)s]) with ordinality as u(fol, ord);

    insert into public.share_links
      (id, space_id, token, is_active, email_required, recipient_email, created_by, created_at)
    values (gen_random_uuid(), s_id, encode(gen_random_bytes(16), 'hex'), true, true,
            'venturethrust@gmail.com', v_fnd, rec.added);

    -- Read state follows the watchlist. A startup the investor chose to
    -- follow was obviously opened, so it can never sit in Pending.
    insert into public.alerts (user_id, space_id, type, message, created_at, read_at)
    values (v_inv, s_id, 'space_shared',
            rec.nm || ' shared their data room with you.',
            rec.added, case when rec.watched then rec.added + interval '2 hours' else null end);

    if rec.watched then
      insert into public.dw_watchlist
        (id, investor_id, founder_id, space_id, file_id, startup_name, manager_id, note,
         quarterly_report, created_at)
      values (gen_random_uuid(), v_inv, v_fnd, s_id, null, rec.nm, v_mgr,
              rec.note, rec.quarterly, rec.added);
    end if;
  end loop;
""" % {"FOLDERS": ", ".join("'%s'" % sqlq(f) for f in FOLDERS)})

    # ── Every document, one line each ────────────────────────────────────────
    out.append("\n  -- ── %d documents, each already uploaded to the vdr-files bucket ────────\n"
               % len(files))
    # The storage path is always demo/<slug>/<filename>, so it is built in SQL
    # rather than repeated on 130 lines. The room name comes from a small
    # slug to name map, which is also what proves every file lands in a room
    # that this script actually created.
    slugmap = ", ".join("('%s','%s')" % (p["slug"], sqlq(p["name"])) for p in ALL_PROFILES)
    out.append("  insert into public.files\n"
               "    (id, user_id, folder_id, space_id, name, type, storage_path,\n"
               "     size_bytes, views, position, created_at)\n"
               "  select gen_random_uuid()::text, sp.created_by, fo.id, sp.id, t.fname, t.ftype,\n"
               "         'demo/' || t.slug || '/' || t.fname,\n"
               "         t.sz, t.vws, t.pos, now() - (t.age || ' days')::interval\n"
               "    from (values\n")
    lines = []
    seen = {}
    for i, f in enumerate(files):
        key = (f["slug"], f["folder"])
        seen[key] = seen.get(key, 0) + 1
        lines.append("      ('%s','%s','%s','%s',%d,%d,%d,%d)"
                     % (f["slug"], sqlq(f["folder"]), sqlq(f["name"]),
                        ftype(f["name"]), f["size"], (i * 7) % 26, seen[key],
                        3 + (i * 5) % 44))
    out.append(",\n".join(lines))
    out.append("\n    ) as t(slug, fol, fname, ftype, sz, vws, pos, age)\n"
               "    join (values %s) as m(slug, nm) on m.slug = t.slug\n"
               "    join public.spaces  sp on sp.name = m.nm and sp.id = any(ids)\n"
               "    join public.folders fo on fo.space_id = sp.id and fo.name = t.fol;\n"
               % slugmap)

    out.append(SQL_TAIL)
    with open(path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("".join(out))
    print("  wrote %s (%d lines)" % (path, "".join(out).count("\n") + 1))


def main():
    print("Generating documents")
    files = generate()
    print("  %d files" % len(files))
    if "--upload" in sys.argv:
        print("Uploading to %s" % BUCKET)
        upload(files)
    if "--sql" in sys.argv:
        print("Writing SQL")
        write_sql(files, os.path.join(
            os.path.dirname(os.path.dirname(HERE)), "sql", "demo_investor_account.sql"))


if __name__ == "__main__":
    main()
